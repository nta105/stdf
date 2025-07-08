import pandas as pd
import re, os
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def get_col_by_testnum(df, testnum: str):
    return next((col for col in df.columns if str(col).startswith(testnum)), None)

def build_serial(msw, lsw, fallback_id):
    try:
        if pd.notna(msw) and pd.notna(lsw):
            return int(msw) << 16 | int(lsw)
        else:
            return f"Unknown_{fallback_id}"
    except (ValueError, TypeError):
        return f"Unknown_{fallback_id}"

def clean_sub_name(sub):
    return re.sub(r'(Coarse Code \d+|Fine Code \d+)', '', sub or "").replace(",", "").strip(', ').strip()

def parse_test_column(col, test_col_pattern):
    match = test_col_pattern.match(col)
    if not match:
        return None
    testnum, test, sub = match.group(1), match.group(2).strip(), match.group(3)
    trim_match = re.search(r'Coarse Code (\d+)', sub or "")
    fine_match = re.search(r'Fine Code (\d+)', sub or "")
    trim = int(trim_match[1]) if trim_match else None
    fine = int(fine_match[1]) if fine_match else None
    sub_clean = clean_sub_name(sub)
    return testnum, test, sub_clean or None, trim, fine

def compute_stats(values):
    if not values:
        return "", "", ""
    avg = sum(values) / len(values)
    stddev = (sum((x - avg) ** 2 for x in values) / len(values)) ** 0.5
    stddev_pct = round(abs(stddev / avg), 2) if abs(avg) > 0 else ""
    return avg, stddev, stddev_pct

def freeze_and_format_excel(output_file):
    wb = load_workbook(output_file)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 14
            col_name = ws[f"{col_letter}1"].value
            if col_name and "%_Diff" in col_name:
                for row in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row}"]
                    if isinstance(cell.value, float):
                        cell.number_format = '0.00'
    wb.save(output_file)

def run_transpose(file_path):
    df = pd.read_excel(file_path)

    test_code_col = "Test_Code"
    if test_code_col not in df.columns:
        raise ValueError("Missing required column: Test_Code.")

    test_codes = df[test_code_col].tolist()
    single_device_mode = all(pd.isna(tc) for tc in test_codes)

    if single_device_mode:
        df[test_code_col] = "T0"
        test_code_set = ["T0"]
    else:
        test_code_set = sorted(set(tc for tc in test_codes if isinstance(tc, str)), key=lambda x: int(re.sub(r'\D', '', x) or 0))

    lsw_col = get_col_by_testnum(df, "40200000")
    msw_col = get_col_by_testnum(df, "40150000")

    serials_raw = [
        build_serial(df.at[i, msw_col] if msw_col else None,
                     df.at[i, lsw_col] if lsw_col else None,
                     i + 1)
        for i in range(len(df))
    ]

    start = next(i for i, col in enumerate(df.columns) if re.match(r"^\d+", str(col)))
    test_cols = df.columns[start:]

    pvin_col = get_col_by_testnum(df, "5030000")
    avin_col = get_col_by_testnum(df, "5035000")
    vddio_col = get_col_by_testnum(df, "5040000")
    temp_col = get_col_by_testnum(df, "5050000")

    serial_order = []
    serial_base_map = {}
    unknown_counter = 1
    for s_raw, tc in zip(serials_raw, test_codes):
        base = f"Unknown_{unknown_counter}" if pd.isna(s_raw) else s_raw
        if pd.isna(s_raw):
            unknown_counter += 1
        full = f"{base}" if single_device_mode else f"{base}_{str(tc)}"
        if full not in serial_order:
            serial_order.append(full)
            serial_base_map[full] = base

    grouped_data = defaultdict(lambda: defaultdict(dict))
    test_col_pattern = re.compile(r"(\d+)\s+([^:]+)(?::(.+))?")

    for i, serial in enumerate(serials_raw):
        serial = f"Unknown_{i+1}" if pd.isna(serial) else serial
        test_code = df.at[i, test_code_col]
        if pd.isna(test_code): continue
        serial_key_prefix = f"{serial}" if single_device_mode else f"{serial}_{test_code}"

        group_key = (df.at[i, pvin_col], df.at[i, avin_col], df.at[i, vddio_col], df.at[i, temp_col])

        for col in test_cols:
            parsed = parse_test_column(col, test_col_pattern)
            if not parsed:
                continue
            test_key = parsed
            value = df.at[i, col]
            existing_keys = [k for k in grouped_data[group_key][test_key] if k.startswith(serial_key_prefix)]
            suffix = f"#{len(existing_keys)+1}" if existing_keys else ""
            full_key = f"{serial_key_prefix}{suffix}"
            grouped_data[group_key][test_key][full_key] = value

    final_rows = []
    for (pvin, avin, vddio, temp), test_dict in grouped_data.items():
        for (testnum, test, sub, trim, fine), serial_values in test_dict.items():
            row = {
                "testnum": testnum, "test": test, "sub": sub, "trim": trim, "fine": fine,
                "PVIN": pvin, "AVIN": avin, "VDDIO": vddio, "TEMP": temp,
            }
            values = [val for val in serial_values.values() if pd.notna(val)]
            for serial_key, val in serial_values.items():
                row[serial_key] = val
            row["Average"], row["StdDev"], row["StdDev%"] = compute_stats(values)
            final_rows.append(row)

    df_out = pd.DataFrame(final_rows)
    meta_cols = ["testnum", "test", "sub", "trim", "fine", "PVIN", "AVIN", "VDDIO", "TEMP"]
    stat_cols = ["Average", "StdDev", "StdDev%"]

    serial_cols = []
    for full_serial in serial_order:
        matching = [col for col in df_out.columns if str(col).startswith(str(full_serial))]
        serial_cols.extend(matching)
    rename_map = {orig: re.sub(r"#\d+$", "", orig) for orig in serial_cols}
    df_out = df_out.rename(columns=rename_map)

    sheet1_df = df_out.drop(columns=stat_cols)
    serial_bases = sorted(set(col.split("_")[0] for col in sheet1_df.columns if "_" in col))

    individual_stat_sheets = {}
    if not single_device_mode:
        for code in test_code_set:
            code_cols = [col for col in df_out.columns if f"_{code}" in col]
            if not code_cols:
                continue
            sheet_df = df_out[meta_cols + code_cols].copy()
            sheet_df["Average"] = sheet_df[code_cols].mean(axis=1, skipna=True)
            sheet_df["StdDev"] = sheet_df[code_cols].std(axis=1, skipna=True)
            sheet_df["StdDev%"] = (sheet_df["StdDev"] / sheet_df["Average"]).abs().round(2)
            individual_stat_sheets[code] = sheet_df

    compare_code_pairs = []
    compare_sheets = {}
    if not single_device_mode and len(test_code_set) > 1:
        compare_code_pairs = [("T0", "T168"), ("T0", "T500"), ("T0", "T1000"), ("T168", "T500"), ("T500", "T1000")]
        for c1, c2 in compare_code_pairs:
            cols = []
            col_names = []
            for base in serial_bases:
                col1 = f"{base}_{c1}"
                col2 = f"{base}_{c2}"
                diff_col = f"{base}_Diff_{c1}_{c2}"
                pct_col = f"{base}_%_Diff_{c1}_{c2}"
                if col1 in df_out.columns and col2 in df_out.columns:
                    col_names.extend([col1, col2, diff_col, pct_col])
                    col1_vals = df_out[col1]
                    col2_vals = df_out[col2]
                    diff_vals = col2_vals - col1_vals
                    pct_vals = 100 * diff_vals / col1_vals.replace(0, pd.NA)
                    pct_vals.replace([float('inf'), float('-inf')], pd.NA, inplace=True)
                    cols.extend([col1_vals, col2_vals, diff_vals, pct_vals])
            if cols:
                data_matrix = pd.concat([df_out[meta_cols]] + cols, axis=1)
                data_matrix.columns = meta_cols + col_names
                compare_sheets[f"{c1}_vs_{c2}"] = data_matrix

    output_file = os.path.join(os.path.dirname(file_path), "Transposed_" + os.path.splitext(os.path.basename(file_path))[0] + ".xlsx")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        sheet1_df.to_excel(writer, index=False, sheet_name="All_NoStats_WithDiff")
        for code, df_sheet in individual_stat_sheets.items():
            df_sheet.to_excel(writer, index=False, sheet_name=f"Only_{code}_Stats")
        for name, df_sheet in compare_sheets.items():
            df_sheet.to_excel(writer, index=False, sheet_name=name)

    freeze_and_format_excel(output_file)
    print(f"Transposed file saved as: {output_file}")
    return output_file
