import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from io import StringIO
import numpy as np
from pystdf.IO import Parser
from pystdf.Writers import TextWriter
import pystdf.V4 as v4

# === Define data types for PTR records ===
RECORD_PTR_DATATYPES = {
    'Record': str, 'LineNum': int, 'SourceFile': str, 'TEST_NUM': int,
    'HEAD_NUM': int, 'SITE_NUM': int, 'TEST_FLG': int, 'PARM_FLG': int,
    'RESULT': float, 'TEST_TXT': str, 'ALARM_ID': float, 'OPT_FLAG': float,
    'RES_SCAL': float, 'LLM_SCAL': float, 'HLM_SCAL': float,
    'LO_LIMIT': float, 'HI_LIMIT': float, 'UNITS': str,
    'C_RESFMT': str, 'C_LLMFMT': str, 'C_HLMFMT': str,
    'LO_SPEC': float, 'HI_SPEC': float,
}

def stdf_to_dfs(file_paths):
    record_dfs = {}
    for file_path in file_paths:
        filename = os.path.basename(file_path)
        p = Parser(inp=open(file_path, 'rb'))
        captured_std_out = StringIO()
        p.addSink(TextWriter(captured_std_out))
        p.parse()
        atdf = captured_std_out.getvalue().split('\n')

        for n, line in enumerate(atdf):
            atdf[n] = line[:4] + str(n) + '|' + filename + '|' + line[4:]

        for record_type in v4.records:
            record_name = record_type.name.split('.')[-1].upper()
            datatypes = RECORD_PTR_DATATYPES if record_name == 'PTR' else None
            curr = '\n'.join([line for line in atdf if line.startswith(record_name)])
            if curr:
                names = [name for name, _ in record_type.fieldMap]
                header_names = ['Record', 'LineNum', 'SourceFile'] + names

                df = pd.read_csv(StringIO(curr), header=None, names=header_names, delimiter='|', dtype=datatypes)
                record_dfs[record_name] = pd.concat([record_dfs.get(record_name, pd.DataFrame()), df])
    return {k: v for k, v in record_dfs.items() if not v.empty}

def process_stdf_file(input_path: str, output_path: str):
    record_dfs = stdf_to_dfs([input_path])
    df_ptr = record_dfs['PTR']
    df_prr = record_dfs['PRR']
    df_mir = record_dfs.get('MIR')
    df_mrr = record_dfs.get('MRR')

    def try_parse_timestamp(val):
        try:
            return datetime.fromtimestamp(int(val)).strftime('%Y-%m-%d %H:%M:%S')
        except:
            return str(val)

    start_time = finish_time = job_name = node_name = ''
    if df_mir is not None and not df_mir.empty:
        row = df_mir.iloc[0]
        start_time = try_parse_timestamp(row.get('START_T', ''))
        job_name = str(row.get('JOB_NAM', ''))
        node_name = str(row.get('NODE_NAM', ''))
    if df_mrr is not None and not df_mrr.empty:
        finish_time = try_parse_timestamp(df_mrr.iloc[0].get('FINISH_T', ''))

    df_ptr = df_ptr[['TEST_NUM', 'TEST_TXT', 'LO_LIMIT', 'HI_LIMIT', 'UNITS', 'RESULT', 'HEAD_NUM', 'SITE_NUM']].copy()
    min_testnum = df_ptr['TEST_NUM'].min()
    device_ids, active_sites = [], {}
    device_counter = 0

    for _, row in df_ptr.iterrows():
        key = (row['HEAD_NUM'], row['SITE_NUM'])
        if row['TEST_NUM'] == min_testnum:
            device_counter += 1
            active_sites[key] = device_counter
        device_ids.append(active_sites.get(key, -1))
    df_ptr['Device_ID'] = device_ids
    unique_devices = df_ptr['Device_ID'].unique()

    serial_map = {}
    for dev_id in unique_devices:
        entry = df_ptr[(df_ptr['Device_ID'] == dev_id) & (df_ptr['TEST_NUM'] == 40250000)]
        if not entry.empty:
            val = entry['RESULT'].values[0]
            serial_map[dev_id] = str(int(val)) if pd.notna(val) else ''
        else:
            serial_map[dev_id] = ''

    ref_device = df_ptr[df_ptr['Device_ID'] == unique_devices[0]][['TEST_NUM', 'TEST_TXT', 'LO_LIMIT', 'HI_LIMIT', 'UNITS']]
    df_merged = ref_device.copy()
    column_name_map = {}
    unnamed_counter = 1

    for dev_id in unique_devices:
        serial = serial_map[dev_id]
        col_name = serial if serial.strip() else f"N/A_{unnamed_counter}"
        if col_name.startswith("N/A_"):
            unnamed_counter += 1
        column_name_map[col_name] = 'N/A' if col_name.startswith('N/A_') else col_name
        dev_df = df_ptr[df_ptr['Device_ID'] == dev_id].groupby(['TEST_NUM', 'TEST_TXT'], as_index=False).agg({'RESULT': 'first'})
        df_merged = pd.merge(df_merged, dev_df.rename(columns={'RESULT': col_name}), on=['TEST_NUM', 'TEST_TXT'], how='left')

    df_prr = df_prr[['SITE_NUM', 'PART_ID', 'SOFT_BIN', 'HARD_BIN', 'PART_FLG']].copy()
    sbin_lookup, hbin_lookup = {}, {}
    for df, key, lookup in [(record_dfs.get('SBR'), 'SBIN_NUM', sbin_lookup), (record_dfs.get('HBR'), 'HBIN_NUM', hbin_lookup)]:
        if df is not None and not df.empty:
            for _, row in df.iterrows():
                num = int(row[key])
                name = str(row.get(f'{key[0]}BIN_NAM', '')).strip()
                if name and name.lower() != 'nan':
                    lookup[num] = name

    def format_bin_label(num, lookup):
        if pd.isna(num): return ''
        num = int(num)
        name = lookup.get(num, '').strip()
        return f"{num} ({name})" if name else str(num)

    prr_info = []
    for i, dev_id in enumerate(unique_devices):
        serial = serial_map[dev_id]
        if i < len(df_prr):
            row = df_prr.iloc[i]
            prr_info.append({
                'serial': serial,
                'meta': [
                    row.get('SITE_NUM', ''),
                    format_bin_label(row.get('SOFT_BIN'), sbin_lookup),
                    format_bin_label(row.get('HARD_BIN'), hbin_lookup),
                    'Pass' if row.get('PART_FLG', 1) == 0 else 'Fail',
                    row.get('PART_ID', '')
                ]
            })
        else:
            prr_info.append({'serial': serial, 'meta': [''] * 5})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Device Summary'

    for i, label in enumerate(['Site Tested', 'Soft Bin #', 'Hard Bin #', 'Pass/Fail', 'Part ID']):
        ws.cell(row=i + 1, column=5, value=label).alignment = Alignment(horizontal='right')

    for col_offset, prr in enumerate(prr_info):
        for row_index, val in enumerate(prr['meta']):
            cell = ws.cell(row=row_index + 1, column=6 + col_offset, value=val)
            if row_index == 3:
                if str(val).lower() == 'pass':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif str(val).lower() == 'fail':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for r_idx, row in enumerate(dataframe_to_rows(df_merged, index=False, header=True), start=6):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=column_name_map.get(val, val) if r_idx == 6 else val)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    for row in ws.iter_rows(min_row=7, min_col=6):
        try:
            lo = float(ws.cell(row=row[0].row, column=3).value)
            hi = float(ws.cell(row=row[0].row, column=4).value)
        except:
            continue
        for cell in row:
            try:
                v = float(cell.value)
                if v < lo or v > hi:
                    cell.fill = red_fill
            except:
                continue

    ws.freeze_panes = 'F7'
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['E'].width = 10
    for i in range(6, 6 + len(prr_info)):
        ws.column_dimensions[get_column_letter(i)].width = 10

    ws['A1'] = 'Start time'; ws['B1'] = start_time
    ws['A2'] = 'Finish time'; ws['B2'] = finish_time
    ws['A3'] = 'Program'; ws['B3'] = job_name
    ws['A4'] = 'Tester'; ws['B4'] = node_name

    import time
    wb.save(output_path)
    time.sleep(0.1)  # wait for file to flush

